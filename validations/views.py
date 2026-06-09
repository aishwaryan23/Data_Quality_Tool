import csv
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST

from .models import ValidationRun, ValidationResult
from mappings.models import Mapping, ColumnMapping, ValidationRule
from accounts.decorators import contributor_or_admin_required

logger = logging.getLogger('validations')


@login_required
def validation_list_view(request):
    """List all validation runs."""
    query = request.GET.get('query', '').strip()
    runs = ValidationRun.objects.select_related('mapping', 'triggered_by').all()
    if query:
        runs = runs.filter(mapping__name__icontains=query)
    runs = runs[:50]
    return render(request, 'validations/list.html', {'runs': runs, 'query': query})


@login_required
def validation_report_view(request, run_id):
    """View detailed report for a validation run."""
    run = get_object_or_404(
        ValidationRun.objects.select_related('mapping', 'triggered_by'),
        id=run_id
    )
    results = run.results.select_related('column_mapping').all()
    return render(request, 'validations/report.html', {
        'run': run,
        'results': results,
    })


@login_required
def validation_progress_view(request, run_id):
    """View validation progress (for active runs)."""
    run = get_object_or_404(ValidationRun, id=run_id)
    return render(request, 'validations/progress.html', {'run': run})


# ─── API Endpoints ───────────────────────────────────────────────────────────

@login_required
def api_validation_progress(request, run_id):
    """AJAX: Get validation progress."""
    run = get_object_or_404(ValidationRun, id=run_id)
    return JsonResponse({
        'status': run.status,
        'progress': run.progress,
        'total': run.total_checks,
        'passed': run.passed_checks,
        'failed': run.failed_checks,
    })


@login_required
@contributor_or_admin_required
@require_POST
def api_trigger_validation(request, mapping_id):
    """AJAX: Manually trigger a validation for a mapping."""
    if hasattr(request.user, 'profile') and request.user.profile.role == 'auditor':
        return JsonResponse({'success': False, 'error': 'Permission denied: Auditor cannot trigger validations.'}, status=403)
    mapping = get_object_or_404(Mapping, id=mapping_id)

    # Extract JSON parameters if present
    import json
    parameters = {}
    if request.content_type == 'application/json':
        try:
            body = json.loads(request.body)
            parameters = body.get('parameters', {})
        except Exception:
            pass
    else:
        param_str = request.POST.get('parameters', '{}')
        try:
            parameters = json.loads(param_str)
        except Exception:
            parameters = {}

    run = ValidationRun.objects.create(
        mapping=mapping,
        triggered_by=request.user,
        trigger_type='manual',
        status='pending',
        parameters=parameters,
    )

    # Try to use Celery, fall back to synchronous
    try:
        from .tasks import run_validation_task
        run_validation_task.delay(run.id)
    except Exception:
        # Fallback: run synchronously
        from .engine import ValidationEngine
        engine = ValidationEngine(run)
        try:
            engine.execute()
            if run.triggered_by:
                from dashboard.models import Notification
                status_text = "Passed" if run.failed_checks == 0 else "Failed"
                Notification.objects.create(
                    user=run.triggered_by,
                    title=f"Validation Run #{run.id} Completed",
                    message=f"Pipeline: {run.mapping.name}\nStatus: {status_text} ({run.passed_checks}/{run.total_checks} checks passed)",
                    level='success' if run.failed_checks == 0 else 'warning'
                )
        except Exception as e:
            logger.error(f"Sync validation failed: {e}")
            if run.triggered_by:
                from dashboard.models import Notification
                Notification.objects.create(
                    user=run.triggered_by,
                    title=f"Validation Run #{run.id} Failed",
                    message=f"Pipeline: {run.mapping.name}\nError: {e}",
                    level='error'
                )

    return JsonResponse({'success': True, 'run_id': run.id})


@login_required
def export_report(request, run_id):
    """Export validation report as Excel file."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    run = get_object_or_404(ValidationRun, id=run_id)
    mapping = run.mapping
    results = run.results.select_related('column_mapping').all()

    src_conn = mapping.source_connection
    col_headers = [
        'Table Name',
        'Column Name',
        'Source Validation Operation(operation choosen)',
        'Target Validation Operation (operation choosen)',
        'Result',
        'Difference'
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    # Write headers
    ws.append(col_headers)

    # Style header row
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Deep Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        res_val = "MATCH" if r.is_match else "MISMATCH"
        ws.append([
            mapping.source_table.upper(),
            r.column_mapping.source_column,
            r.source_op_display,
            r.target_op_display,
            res_val,
            r.difference
        ])

    # Formatting columns
    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if col in [1, 2, 3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col == 5:
                if cell.value == "MATCH":
                    cell.font = Font(name="Calibri", size=11, bold=True, color="15803D") # Green
                else:
                    cell.font = Font(name="Calibri", size=11, bold=True, color="B91C1C") # Red

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    val_date = run.created_at.strftime('%Y-%m-%d')
    filename = f"{mapping.source_table.upper()}_{val_date}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    # Log report generation audit
    try:
        from logs.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action=f"Report Generated: {mapping.source_table} on {val_date}",
            entity_type="ValidationRun",
            entity_id=run.id,
            level="success",
        )
    except Exception:
        pass

    return response


def get_datatype_category(type_str, name_str=''):
    t = str(type_str or '').upper()
    n = str(name_str or '').upper()
    if any(x in t for x in ('INT', 'BIGINT', 'SMALLINT', 'TINYINT', 'NUMERIC', 'DECIMAL', 'FLOAT', 'DOUBLE', 'REAL', 'NUMBER')):
        return 'INTEGER'
    elif (any(x in t for x in ('DATE', 'TIME', 'TIMESTAMP')) or 
          any(x in n for x in ('DATE', 'TIME', 'TIMESTAMP')) or 
          n.endswith('_AT') or n.endswith('_ON') or n == 'AT' or 'DT' in n):
        return 'DATE'
    elif any(x in t for x in ('BOOL', 'BOOLEAN')):
        return 'BOOLEAN'
    else:
        return 'VARCHAR'

def get_applicable_operations(category):
    if category == 'INTEGER':
        return ['null_check', 'sum', 'avg', 'min', 'max', 'range_check', 'duplicate_check', 'unique_check', 'distinct_count', 'count', 'row_count', 'data_type_check']
    elif category == 'DATE':
        return ['min_date', 'max_date', 'null_check', 'duplicate_check', 'unique_check', 'distinct_count', 'count', 'row_count']
    elif category == 'BOOLEAN':
        return ['null_check', 'count', 'row_count', 'duplicate_check', 'unique_check', 'distinct_count']
    else: # VARCHAR
        return ['null_check', 'length_sum_check', 'sum_length', 'regex_check', 'duplicate_check', 'unique_check', 'distinct_count', 'row_count', 'count', 'case_insensitive_check', 'trim_check', 'contains_check', 'pattern_match', 'data_type_check']

@login_required
@contributor_or_admin_required
@require_POST
def validation_delete_view(request, run_id):
    """Delete a validation run."""
    run = get_object_or_404(ValidationRun, id=run_id)
    run.delete()
    messages.success(request, f"Validation Run #{run_id} deleted successfully.")
    return redirect('validations:list')


@login_required
@contributor_or_admin_required
def quick_validate_view(request):
    """Create a quick mapping and trigger validation from the dashboard."""
    if request.method == 'POST':
        try:
            source_conn_id = request.POST.get('source_connection')
            source_schema = request.POST.get('source_schema', '')
            source_table = request.POST.get('source_table', '')
            target_conn_id = request.POST.get('target_connection')
            target_schema = request.POST.get('target_schema', '')
            target_table = request.POST.get('target_table', '')
            
            # Resolving Source Date Filters
            source_date_column = request.POST.get('source_date_column', '')
            source_date_filter_type = request.POST.get('source_date_filter_type', 'none')
            if not source_date_column:
                source_date_filter_type = 'none'
            source_date_filter_start = None
            source_date_filter_end = None
            source_date_value_type = request.POST.get('source_date_value_type', 'calendar')
            source_date_relative_operator = request.POST.get('source_date_relative_operator', '+')
            try:
                source_date_relative_value = int(request.POST.get('source_date_relative_value', 0) or 0)
            except (ValueError, TypeError):
                source_date_relative_value = 0
            source_date_operator = request.POST.get('source_date_operator', '=')

            if source_date_filter_type == 'specific':
                if source_date_value_type == 'calendar':
                    source_date_single = request.POST.get('source_date_single')
                    if source_date_single:
                        source_date_filter_start = source_date_single
                        source_date_filter_end = source_date_single
            elif source_date_filter_type == 'range':
                source_date_filter_start = request.POST.get('source_date_filter_start') or None
                source_date_filter_end = request.POST.get('source_date_filter_end') or None

            # Resolving Target Date Filters
            target_date_column = request.POST.get('target_date_column', '')
            target_date_filter_type = request.POST.get('target_date_filter_type', 'none')
            if not target_date_column:
                target_date_filter_type = 'none'
            target_date_filter_start = None
            target_date_filter_end = None
            target_date_value_type = request.POST.get('target_date_value_type', 'calendar')
            target_date_relative_operator = request.POST.get('target_date_relative_operator', '+')
            try:
                target_date_relative_value = int(request.POST.get('target_date_relative_value', 0) or 0)
            except (ValueError, TypeError):
                target_date_relative_value = 0
            target_date_operator = request.POST.get('target_date_operator', '=')

            if target_date_filter_type == 'specific':
                if target_date_value_type == 'calendar':
                    target_date_single = request.POST.get('target_date_single')
                    if target_date_single:
                        target_date_filter_start = target_date_single
                        target_date_filter_end = target_date_single
            elif target_date_filter_type == 'range':
                target_date_filter_start = request.POST.get('target_date_filter_start') or None
                target_date_filter_end = request.POST.get('target_date_filter_end') or None

            # Create a quick mapping
            from django.utils import timezone
            now_str = timezone.now().strftime('%Y-%m-%d %H:%M')
            quick_name = f"Quick Validate: {source_table} -> {target_table} ({now_str})".strip()
            
            mapping_data = {
                'name': quick_name,
                'description': "Triggered from Dashboard Quick Workspace",
                'source_connection_id': source_conn_id,
                'source_schema': source_schema,
                'source_table': source_table,
                'target_connection_id': target_conn_id,
                'target_schema': target_schema,
                'target_table': target_table,
                'created_by': request.user,
                'source_date_column': source_date_column,
                'source_date_filter_type': source_date_filter_type,
                'source_date_filter_start': source_date_filter_start,
                'source_date_filter_end': source_date_filter_end,
                'source_date_operator': source_date_operator,
                'target_date_column': target_date_column,
                'target_date_filter_type': target_date_filter_type,
                'target_date_filter_start': target_date_filter_start,
                'target_date_filter_end': target_date_filter_end,
                'target_date_operator': target_date_operator,
                # Stale fields that might be submitted from client or UI
                'source_date_value_type': source_date_value_type,
                'source_date_relative_operator': source_date_relative_operator,
                'source_date_relative_value': source_date_relative_value,
                'target_date_value_type': target_date_value_type,
                'target_date_relative_operator': target_date_relative_operator,
                'target_date_relative_value': target_date_relative_value,
            }

            # Dynamic fields verification and defensive logging
            model_fields = set()
            for f in Mapping._meta.get_fields():
                model_fields.add(f.name)
                if hasattr(f, 'attname'):
                    model_fields.add(f.attname)

            rejected_fields = {}
            for key in list(mapping_data.keys()):
                if key not in model_fields:
                    rejected_fields[key] = mapping_data[key]
                    logger.warning(
                        f"Field mismatch: Submitted field '{key}' is not a valid attribute of the Mapping model. "
                        f"Removing from parameters list. "
                        f"Model fields: {sorted(list(model_fields))}"
                    )
                    del mapping_data[key]

            try:
                mapping = Mapping.objects.create(**mapping_data)
            except Exception as e:
                logger.error(
                    f"Quick validation creation error: {e}. "
                    f"Submitted keys: {list(mapping_data.keys())}. "
                    f"Rejected data: {rejected_fields}. "
                    f"Model fields: {sorted(list(model_fields))}"
                )
                raise

            try:
                from dashboard.models import FormDraft
                FormDraft.objects.filter(user=request.user, page_key='quick_validate', status='draft').update(status='completed')
            except Exception:
                pass

            # Read columns json
            column_data = request.POST.get('column_mappings_json', '[]')
            import json
            try:
                columns = json.loads(column_data)
            except json.JSONDecodeError:
                columns = []

            # Fallback values if JSON is empty but old format is present
            source_cols = request.POST.getlist('source_columns')
            target_cols = request.POST.getlist('target_columns')
            selected_ops = request.POST.getlist('operations')

            # If "__all__" is passed in columns mapping
            if (columns and columns[0].get('source_column') == '__all__') or ('__all__' in source_cols or '__all__' in target_cols):
                user_selected_ops = []
                if columns and columns[0].get('source_column') == '__all__':
                    user_selected_ops = columns[0].get('operations', [])
                elif selected_ops:
                    user_selected_ops = selected_ops

                from connections.connector import ConnectorEngine
                source_conn = mapping.source_connection
                target_conn = mapping.target_connection
                source_engine = ConnectorEngine(source_conn)
                target_engine = ConnectorEngine(target_conn)
                
                src_all_cols = source_engine.get_columns(source_schema if source_schema != 'file' else None, source_table)
                tgt_all_cols = target_engine.get_columns(target_schema if target_schema != 'file' else None, target_table)
                
                expanded_columns = []
                for s_col in src_all_cols:
                    matched_t = next((t_col for t_col in tgt_all_cols if t_col['name'].lower() == s_col['name'].lower()), None)
                    if matched_t:
                        s_cat = get_datatype_category(s_col['type'], s_col['name'])
                        all_ops = get_applicable_operations(s_cat)
                        if user_selected_ops:
                            ops = [op for op in all_ops if op in user_selected_ops]
                        else:
                            ops = all_ops
                        expanded_columns.append({
                            'source_column': s_col['name'],
                            'source_datatype': s_col['type'],
                            'target_column': matched_t['name'],
                            'target_datatype': matched_t['type'],
                            'operations': ops
                        })
                columns = expanded_columns
            
            if not columns and source_cols and target_cols:
                # Old manual pairing fallback
                max_len = max(len(source_cols), len(target_cols))
                for i in range(max_len):
                    src = source_cols[i] if i < len(source_cols) else ''
                    tgt = target_cols[i] if i < len(target_cols) else ''
                    if src and tgt:
                        columns.append({
                            'source_column': src,
                            'source_datatype': 'unknown',
                            'target_column': tgt,
                            'target_datatype': 'unknown',
                            'operations': selected_ops
                        })

            # Create column mappings and rules
            for col in columns:
                s_col = col.get('source_column', '')
                t_col = col.get('target_column', '')
                s_type = col.get('source_datatype', 'unknown')
                t_type = col.get('target_datatype', 'unknown')
                
                col_mapping = ColumnMapping.objects.create(
                    mapping=mapping,
                    source_column=s_col,
                    source_datatype=s_type,
                    target_column=t_col,
                    target_datatype=t_type,
                )
                
                for op in col.get('operations', []):
                    ValidationRule.objects.create(
                        column_mapping=col_mapping,
                        operation=op,
                    )
            
            # Create Validation Run
            param_str = request.POST.get('parameters', '{}')
            try:
                parameters = json.loads(param_str)
            except Exception:
                parameters = {}

            run = ValidationRun.objects.create(
                mapping=mapping,
                triggered_by=request.user,
                trigger_type='manual',
                status='pending',
                source_date_filter_start=source_date_filter_start,
                source_date_filter_end=source_date_filter_end,
                target_date_filter_start=target_date_filter_start,
                target_date_filter_end=target_date_filter_end,
                parameters=parameters,
            )
            
            # Execute validation run (eager or async)
            try:
                from .tasks import run_validation_task
                run_validation_task.delay(run.id)
            except Exception:
                from .engine import ValidationEngine
                engine = ValidationEngine(run)
                try:
                    engine.execute()
                    if run.triggered_by:
                        from dashboard.models import Notification
                        status_text = "Passed" if run.failed_checks == 0 else "Failed"
                        Notification.objects.create(
                            user=run.triggered_by,
                            title=f"Validation Run #{run.id} Completed",
                            message=f"Pipeline: {run.mapping.name}\nStatus: {status_text} ({run.passed_checks}/{run.total_checks} checks passed)",
                            level='success' if run.failed_checks == 0 else 'warning'
                        )
                except Exception as e:
                    logger.error(f"Sync quick validation failed: {e}")
                    if run.triggered_by:
                        from dashboard.models import Notification
                        Notification.objects.create(
                            user=run.triggered_by,
                            title=f"Validation Run #{run.id} Failed",
                            message=f"Pipeline: {run.mapping.name}\nError: {e}",
                            level='error'
                        )
            
            messages.success(request, 'Quick validation triggered successfully!')
            return redirect('validations:progress', run_id=run.id)
            
        except Exception as e:
            logger.error(f"Quick validation creation error: {e}")
            messages.error(request, f"Failed to run quick validation: {str(e)}")
            return redirect('dashboard:index')
            
    return redirect('dashboard:index')


@login_required
def api_mapping_rules_metadata(request, mapping_id):
    """AJAX: Get validation rules requiring user parameters."""
    mapping = get_object_or_404(Mapping, id=mapping_id)
    rules_needing_params = []
    
    column_mappings = mapping.column_mappings.all()
    for cm in column_mappings:
        for rule in cm.rules.filter(is_active=True):
            if rule.operation in ('contains_check', 'starts_with_check', 'ends_with_check', 'pattern_match'):
                rules_needing_params.append({
                    'id': rule.id,
                    'column': cm.source_column,
                    'operation': rule.operation,
                    'operation_display': rule.get_operation_display(),
                })
                
    return JsonResponse({
        'requires_parameters': len(rules_needing_params) > 0,
        'rules': rules_needing_params
    })
